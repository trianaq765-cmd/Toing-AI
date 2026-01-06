"""
Excel Engine - Baca, tulis, manipulasi, analisis file Excel
Support: .xlsx, .xls
Format: Indonesia (Rupiah, tanggal DD/MM/YYYY)
"""

import asyncio
import logging
from pathlib import Path
from typing import Optional, Dict, Any, List, Union, Tuple
from dataclasses import dataclass
from datetime import datetime
import json

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import xlsxwriter

from config import (
    TEMP_DIR, CURRENCY_SYMBOL, DATE_FORMAT,
    MAX_EXCEL_ROWS, MAX_EXCEL_COLS, COLORS
)

logger = logging.getLogger("office_bot.excel")

# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class ExcelError:
    """Excel error information"""
    cell: str
    error_type: str
    formula: str
    message: str

@dataclass
class ExcelStats:
    """Excel file statistics"""
    total_sheets: int
    total_rows: int
    total_columns: int
    has_formulas: bool
    has_errors: bool
    errors: List[ExcelError]
    file_size: int

@dataclass
class CellStyle:
    """Cell styling"""
    bold: bool = False
    italic: bool = False
    font_size: int = 11
    font_color: str = "000000"
    bg_color: Optional[str] = None
    align_h: str = "general"  # left, center, right
    align_v: str = "bottom"   # top, center, bottom
    border: bool = False
    number_format: Optional[str] = None

# =============================================================================
# EXCEL ENGINE CLASS
# =============================================================================

class ExcelEngine:
    """
    Excel manipulation engine with Indonesian formatting
    """
    
    def __init__(self):
        self.temp_dir = TEMP_DIR
        self.temp_dir.mkdir(exist_ok=True)
        
        # Error types yang bisa dideteksi
        self.error_types = [
            "#DIV/0!", "#N/A", "#NAME?", "#NULL!", 
            "#NUM!", "#REF!", "#VALUE!", "#GETTING_DATA"
        ]
        
        logger.info("Excel Engine initialized")
    
    # =========================================================================
    # FILE OPERATIONS
    # =========================================================================
    
    async def read_excel(
        self, 
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Baca file Excel dan return data + metadata
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File tidak ditemukan: {file_path}")
        
        # Check file size
        file_size = file_path.stat().st_size
        if file_size > 25 * 1024 * 1024:  # 25MB limit
            raise ValueError("File terlalu besar (max 25MB)")
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=False)
            
            # Get sheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan")
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Extract data
            data = await self._extract_sheet_data(ws)
            
            # Get statistics
            stats = await self._get_workbook_stats(wb)
            
            # Detect errors
            errors = await self._detect_errors(ws)
            
            wb.close()
            
            return {
                "success": True,
                "file_name": file_path.name,
                "file_size": file_size,
                "sheets": wb.sheetnames,
                "active_sheet": ws.title,
                "data": data,
                "stats": stats,
                "errors": errors
            }
            
        except Exception as e:
            logger.error(f"Error reading Excel: {e}")
            raise
    
    async def create_excel(
        self,
        structure: Dict[str, Any],
        output_path: Optional[Union[str, Path]] = None
    ) -> Path:
        """
        Buat file Excel dari struktur JSON
        
        Structure format:
        {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["Col1", "Col2"],
                    "data": [[val1, val2], ...],
                    "formulas": {"A10": "=SUM(A1:A9)"},
                    "column_widths": {"A": 15, "B": 20},
                    "formatting": {...}
                }
            ]
        }
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.temp_dir / f"excel_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)
        
        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            sheets_data = structure.get("sheets", [])
            if not sheets_data:
                raise ValueError("Tidak ada data sheet dalam struktur")
            
            for sheet_config in sheets_data:
                await self._create_sheet(wb, sheet_config)
            
            # Save workbook
            wb.save(output_path)
            wb.close()
            
            logger.info(f"Excel created: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating Excel: {e}")
            raise
    
    async def modify_excel(
        self,
        file_path: Union[str, Path],
        modifications: Dict[str, Any],
        output_path: Optional[Union[str, Path]] = None
    ) -> Path:
        """
        Modify existing Excel file
        
        Modifications format:
        {
            "sheet": "Sheet1",
            "updates": {
                "A1": "New Value",
                "B2": "=SUM(A1:A10)"
            },
            "delete_rows": [5, 6],
            "insert_rows": [(3, {"A": "data", "B": "data2"})],
            "apply_style": {"A1:B10": {...}}
        }
        """
        file_path = Path(file_path)
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.temp_dir / f"modified_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)
        
        try:
            wb = load_workbook(file_path)
            
            sheet_name = modifications.get("sheet")
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Apply updates
            if "updates" in modifications:
                await self._apply_cell_updates(ws, modifications["updates"])
            
            # Delete rows
            if "delete_rows" in modifications:
                for row in sorted(modifications["delete_rows"], reverse=True):
                    ws.delete_rows(row)
            
            # Insert rows
            if "insert_rows" in modifications:
                for row_idx, row_data in modifications["insert_rows"]:
                    ws.insert_rows(row_idx)
                    for col, value in row_data.items():
                        ws[f"{col}{row_idx}"] = value
            
            # Apply styling
            if "apply_style" in modifications:
                for cell_range, style in modifications["apply_style"].items():
                    await self._apply_style(ws, cell_range, style)
            
            wb.save(output_path)
            wb.close()
            
            logger.info(f"Excel modified: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error modifying Excel: {e}")
            raise
    
    async def fix_excel_errors(
        self,
        file_path: Union[str, Path],
        auto_fix: bool = True
    ) -> Tuple[Path, List[ExcelError]]:
        """
        Detect and fix Excel errors
        """
        file_path = Path(file_path)
        
        # Read and detect errors
        excel_data = await self.read_excel(file_path)
        errors = excel_data["errors"]
        
        if not errors:
            logger.info("No errors found in Excel")
            return file_path, []
        
        if not auto_fix:
            return file_path, errors
        
        # Auto fix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = self.temp_dir / f"fixed_{timestamp}.xlsx"
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            fixed_count = 0
            for error in errors:
                if await self._try_fix_error(ws, error):
                    fixed_count += 1
            
            wb.save(output_path)
            wb.close()
            
            logger.info(f"Fixed {fixed_count}/{len(errors)} errors")
            return output_path, errors
            
        except Exception as e:
            logger.error(f"Error fixing Excel: {e}")
            raise
    
    # =========================================================================
    # SHEET OPERATIONS
    # =========================================================================
    
    async def _create_sheet(self, wb: Workbook, config: Dict[str, Any]):
        """Create a sheet with data and formatting"""
        sheet_name = config.get("name", "Sheet1")
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        headers = config.get("headers", [])
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # Header style
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
        
        # Add data
        data = config.get("data", [])
        start_row = 2 if headers else 1
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                # Check if value is formula
                if isinstance(value, str) and value.startswith("="):
                    ws.cell(row=row_idx, column=col_idx).value = value
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add additional formulas
        formulas = config.get("formulas", {})
        for cell_addr, formula in formulas.items():
            ws[cell_addr] = formula
        
        # Set column widths
        column_widths = config.get("column_widths", {})
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Apply formatting
        formatting = config.get("formatting", {})
        await self._apply_formatting(ws, formatting)
    
    async def _extract_sheet_data(self, ws: Worksheet) -> List[List[Any]]:
        """Extract data from worksheet"""
        data = []
        max_row = min(ws.max_row, MAX_EXCEL_ROWS)
        max_col = min(ws.max_column, MAX_EXCEL_COLS)
        
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            row_data = []
            for cell in row:
                if cell.value is None:
                    row_data.append(None)
                elif cell.data_type == 'f':  # Formula
                    row_data.append(f"={cell.value}")
                else:
                    row_data.append(cell.value)
            data.append(row_data)
        
        return data
    
    # =========================================================================
    # ERROR DETECTION & FIXING
    # =========================================================================
    
    async def _detect_errors(self, ws: Worksheet) -> List[ExcelError]:
        """Detect errors in worksheet"""
        errors = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in self.error_types:
                    errors.append(ExcelError(
                        cell=cell.coordinate,
                        error_type=cell.value,
                        formula=str(cell.value) if cell.data_type == 'f' else "",
                        message=self._get_error_message(cell.value)
                    ))
        
        return errors
    
    def _get_error_message(self, error_type: str) -> str:
        """Get error description"""
        messages = {
            "#DIV/0!": "Pembagian dengan nol",
            "#N/A": "Nilai tidak tersedia",
            "#NAME?": "Nama fungsi tidak dikenali",
            "#NULL!": "Interseksi range tidak valid",
            "#NUM!": "Nilai numerik tidak valid",
            "#REF!": "Referensi cell tidak valid",
            "#VALUE!": "Tipe data tidak sesuai",
            "#GETTING_DATA": "Mengambil data..."
        }
        return messages.get(error_type, "Error tidak dikenal")
    
    async def _try_fix_error(self, ws: Worksheet, error: ExcelError) -> bool:
        """Try to auto-fix common errors"""
        cell = ws[error.cell]
        
        # Fix #DIV/0! dengan IFERROR
        if error.error_type == "#DIV/0!" and cell.data_type == 'f':
            original_formula = cell.value
            cell.value = f'=IFERROR({original_formula}, 0)'
            return True
        
        # Fix #N/A dengan IFERROR
        if error.error_type == "#N/A" and cell.data_type == 'f':
            original_formula = cell.value
            cell.value = f'=IFERROR({original_formula}, "-")'
            return True
        
        return False
    
    # =========================================================================
    # STYLING
    # =========================================================================
    
    async def _apply_formatting(self, ws: Worksheet, formatting: Dict[str, Any]):
        """Apply formatting to worksheet"""
        
        # Currency columns
        currency_cols = formatting.get("currency_columns", [])
        for col in currency_cols:
            for cell in ws[col]:
                if cell.row > 1 and cell.value is not None:
                    cell.number_format = f'"{CURRENCY_SYMBOL}" #,##0'
        
        # Percentage columns
        percentage_cols = formatting.get("percentage_columns", [])
        for col in percentage_cols:
            for cell in ws[col]:
                if cell.row > 1 and cell.value is not None:
                    cell.number_format = '0.00%'
        
        # Date columns
        date_cols = formatting.get("date_columns", [])
        for col in date_cols:
            for cell in ws[col]:
                if cell.row > 1 and cell.value is not None:
                    cell.number_format = 'DD/MM/YYYY'
        
        # Auto-fit columns
        if formatting.get("auto_fit", False):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    async def _apply_style(self, ws: Worksheet, cell_range: str, style: Dict[str, Any]):
        """Apply style to cell range"""
        
        # Parse range (e.g., "A1:B10" or "A1")
        if ":" in cell_range:
            cells = ws[cell_range]
        else:
            cells = [[ws[cell_range]]]
        
        # Apply to all cells in range
        for row in cells:
            for cell in row if isinstance(row, tuple) else [row]:
                if style.get("bold"):
                    cell.font = Font(bold=True)
                
                if style.get("bg_color"):
                    cell.fill = PatternFill(
                        start_color=style["bg_color"],
                        end_color=style["bg_color"],
                        fill_type="solid"
                    )
                
                if style.get("align"):
                    cell.alignment = Alignment(
                        horizontal=style.get("align", "general")
                    )
                
                if style.get("border"):
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
    
    async def _apply_cell_updates(self, ws: Worksheet, updates: Dict[str, Any]):
        """Apply cell updates"""
        for cell_addr, value in updates.items():
            ws[cell_addr] = value
    
    # =========================================================================
    # STATISTICS
    # =========================================================================
    
    async def _get_workbook_stats(self, wb: Workbook) -> ExcelStats:
        """Get workbook statistics"""
        total_rows = 0
        total_columns = 0
        has_formulas = False
        
        for sheet in wb.worksheets:
            total_rows += sheet.max_row
            total_columns = max(total_columns, sheet.max_column)
            
            # Check for formulas
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        has_formulas = True
                        break
                if has_formulas:
                    break
        
        return ExcelStats(
            total_sheets=len(wb.worksheets),
            total_rows=total_rows,
            total_columns=total_columns,
            has_formulas=has_formulas,
            has_errors=False,  # Will be updated by caller
            errors=[],
            file_size=0  # Will be updated by caller
        )
    
    # =========================================================================
    # DATA CONVERSION
    # =========================================================================
    
    async def excel_to_dataframe(
        self, 
        file_path: Union[str, Path],
        sheet_name: Optional[str] = None
    ) -> pd.DataFrame:
        """Convert Excel to Pandas DataFrame"""
        file_path = Path(file_path)
        
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name or 0,
                engine='openpyxl'
            )
            return df
        except Exception as e:
            logger.error(f"Error converting to DataFrame: {e}")
            raise
    
    async def dataframe_to_excel(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        sheet_name: str = "Sheet1",
        apply_formatting: bool = True
    ) -> Path:
        """Convert Pandas DataFrame to Excel"""
        output_path = Path(output_path)
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                if apply_formatting:
                    wb = writer.book
                    ws = wb[sheet_name]
                    
                    # Format header
                    for cell in ws[1]:
                        cell.font = Font(bold=True, size=12)
                        cell.alignment = Alignment(horizontal="center")
                        cell.fill = PatternFill(
                            start_color="366092",
                            end_color="366092",
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
            
            logger.info(f"DataFrame saved to Excel: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error converting DataFrame to Excel: {e}")
            raise
    
    # =========================================================================
    # UTILITIES
    # =========================================================================
    
    def get_cell_value(self, ws: Worksheet, cell: str) -> Any:
        """Get cell value"""
        return ws[cell].value
    
    def set_cell_value(self, ws: Worksheet, cell: str, value: Any):
        """Set cell value"""
        ws[cell] = value
    
    async def get_column_data(
        self, 
        file_path: Union[str, Path],
        column: str,
        sheet_name: Optional[str] = None
    ) -> List[Any]:
        """Get all data from a column"""
        data = await self.read_excel(file_path, sheet_name)
        
        # Determine column index
        col_idx = ord(column.upper()) - ord('A')
        
        column_data = []
        for row in data["data"]:
            if col_idx < len(row):
                column_data.append(row[col_idx])
        
        return column_data
    
    async def create_template(self, template_type: str) -> Path:
        """
        Create predefined Excel templates
        
        Types: invoice, payroll, expense_report, attendance, etc.
        """
        templates = {
            "invoice": await self._template_invoice(),
            "payroll": await self._template_payroll(),
            "expense": await self._template_expense(),
            "attendance": await self._template_attendance(),
        }
        
        if template_type not in templates:
            raise ValueError(f"Template '{template_type}' tidak tersedia")
        
        return await self.create_excel(templates[template_type])
    
    async def _template_invoice(self) -> Dict[str, Any]:
        """Invoice template"""
        return {
            "sheets": [{
                "name": "Invoice",
                "headers": ["No", "Deskripsi", "Qty", "Harga Satuan", "Total"],
                "data": [
                    [1, "", "", "", "=C2*D2"],
                ],
                "formulas": {
                    "E10": "=SUM(E2:E9)",
                    "E11": "=E10*0.11",  # PPN 11%
                    "E12": "=E10+E11"
                },
                "column_widths": {"A": 5, "B": 40, "C": 10, "D": 15, "E": 15},
                "formatting": {
                    "currency_columns": ["D", "E"],
                    "auto_fit": False
                }
            }]
        }
    
    async def _template_payroll(self) -> Dict[str, Any]:
        """Payroll template"""
        return {
            "sheets": [{
                "name": "Payroll",
                "headers": [
                    "NIK", "Nama", "Gaji Pokok", "Tunjangan",
                    "Total Pendapatan", "BPJS Kesehatan (1%)",
                    "BPJS TK (2%)", "PPh 21", "Total Potongan", "Gaji Bersih"
                ],
                "data": [],
                "column_widths": {
                    "A": 12, "B": 25, "C": 15, "D": 15,
                    "E": 18, "F": 18, "G": 15, "H": 15, "I": 18, "J": 18
                },
                "formatting": {
                    "currency_columns": ["C", "D", "E", "F", "G", "H", "I", "J"]
                }
            }]
        }
    
    async def _template_expense(self) -> Dict[str, Any]:
        """Expense report template"""
        return {
            "sheets": [{
                "name": "Expense Report",
                "headers": ["Tanggal", "Kategori", "Deskripsi", "Jumlah", "Keterangan"],
                "data": [],
                "column_widths": {"A": 12, "B": 20, "C": 40, "D": 15, "E": 30},
                "formatting": {
                    "date_columns": ["A"],
                    "currency_columns": ["D"]
                }
            }]
        }
    
    async def _template_attendance(self) -> Dict[str, Any]:
        """Attendance template"""
        return {
            "sheets": [{
                "name": "Absensi",
                "headers": [
                    "NIK", "Nama", "Tanggal", "Jam Masuk",
                    "Jam Keluar", "Status", "Keterangan"
                ],
                "data": [],
                "column_widths": {
                    "A": 12, "B": 25, "C": 12, "D": 12,
                    "E": 12, "F": 15, "G": 30
                },
                "formatting": {
                    "date_columns": ["C"]
                }
            }]
  }
